import os
import pytest
import re
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.service import Service

@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.maximize_window()
    yield driver
    driver.quit()

# @pytest.fixture
def excel_data():
    excel_file = 'AD_datasheet.xlsx'
    sheet_name = 'Sheet1'
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheet_name]
    nRows = sheet.max_row

    data = []
    for row in range(2, nRows + 1):
        subject = sheet.cell(row, 2).value
        message = sheet.cell(row, 3).value
        fileList = sheet.cell(row, 4).value
        expectedResult = sheet.cell(row, 5).value

        # Process discussion's subject
        if subject == "null":
            subject = ""

        # Process discussion's message
        if message == "null":
            message = ""

        # Process filename list
        if fileList == "null":
            files = []
        else:
            files = re.split("; ", fileList)
                    
        data.append((row, subject, message, files, expectedResult))

    workbook.close()
    return data

class TestAddDiscussion:
    def update_excel(self, row, result):
        excel_file = 'AD_datasheet.xlsx'
        sheet_name = 'Sheet1'
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=6, value=result)
        workbook.save(excel_file)
        workbook.close()
    
    @pytest.mark.parametrize("row, subject, message, files, expectedResult", excel_data())
    def test_addDiscussion(self, driver, row, subject, message, files, expectedResult):
        # Test logic using the provided WebDriver instance (driver)
        driver.get("https://school.moodledemo.net/login/index.php?loginredirect=1")
        driver.maximize_window()

        # Login as Teacher
        driver.find_element(By.ID, "username").click()
        driver.find_element(By.ID, "username").send_keys("teacher")
        driver.find_element(By.ID, "password").click()
        driver.find_element(By.ID, "password").send_keys("moodle")
        driver.find_element(By.ID, "loginbtn").click()

        # Go to Add discussion topic
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.LINK_TEXT,"Moodle and Mountaineering")))
        driver.find_element(By.LINK_TEXT,"Moodle and Mountaineering").click()
        link_course = driver.current_url
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#module-944 > div > div.activity-grid > div.activity-name-area.activity-instance.d-flex.flex-column.mr-2")))
        driver.find_element(By.CSS_SELECTOR, "#module-944 > div > div.activity-grid > div.activity-name-area.activity-instance.d-flex.flex-column.mr-2").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.LINK_TEXT, "Add discussion topic")))
        driver.find_element(By.LINK_TEXT, "Add discussion topic").click()
        
        # Data-driven testing - Subject
        driver.find_element(By.ID, "id_subject_label").click()
        driver.find_element(By.ID, "id_subject_label").send_keys(subject)
        
        # Data-driven testing - Message
        driver.find_element(By.ID, "id_message_label").click()
        driver.find_element(By.ID, "id_message_label").send_keys(message)

        # Data-driven testing - File upload
        if len(files) != "0":
            driver.find_element(By.ID, "id_advancedadddiscussion").click()
            driver.find_element(By.ID, "yui_3_18_1_1_1715445250175_571").click()
            for i in range(0, len(files)):
                driver.find_element(By.ID, "yui_3_18_1_1_1715445250175_571").send_keys(os.getcwd() + files[i])
            
        # Submit    
        driver.find_element(By.ID, "id_submitbutton").click()
        
        if expectedResult == "Add new discussion successfully":
            try:
                driver.current_url == link_course
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")
            
        elif expectedResult == "Alert: Invalid Subject":
            try:
                driver.find_element(By.ID,'id_error_subject')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")
                
        elif expectedResult == "Alert: Invalid Message":
            try:
                driver.find_element(By.ID,'id_error_message')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")

        elif expectedResult == "Alert: File size is too big":
            try:
                driver.find_element(By.ID,'id_error_timelimit')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")

        driver.find_element(By.ID, "user-menu-toggle").click()
        driver.find_element(By.LINK_TEXT, "Log out").click()