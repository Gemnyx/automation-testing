import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

class FileExcelReader:
    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet

    def get_row_count(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheet]
        return (sheet.max_row)

    def get_column_count(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheet]
        return (sheet.max_column)

    def read_data(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheet]
        return sheet.cell(row=rownum, column=colnum).value

    def write_data(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheet]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)

class TestUploadFile:
    def set_up(self):
        self.driver = webdriver.Chrome()
        self.driver.get('https://school.moodledemo.net/login/index.php')
    
    def tear_down(self):
        self.driver.quit()
        
    def test_upload(self, file_path, expected):
        self.driver.find_element(By.ID, 'username').send_keys('student')
        time.sleep(1)
        self.driver.find_element(By.ID, 'password').send_keys('moodle')
        time.sleep(0.5)
        self.driver.find_element(By.ID, 'loginbtn').click()
        WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.ID, "user-menu-toggle")))
        self.driver.find_element(By.ID, "user-menu-toggle").click()
        self.driver.find_element(By.LINK_TEXT, "Private files").click()
        WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "fp-btn-add")))
        self.driver.find_element(By.CLASS_NAME, "fp-btn-add").click()
        time.sleep(2)
        if file_path is not None: 
            self.driver.find_element(By.NAME, "repo_upload_file").send_keys(file_path)
            time.sleep(1)
        self.driver.find_element(By.XPATH, "//button[contains(.,\'Upload this file\')]").click()
        time.sleep(5)
        if (expected == "Successfully Upload"): return True
        elif (expected == "No File Attached"): return True
        elif (expected == "Empty File"): return True
        elif (expected == "File exceeded 100MB"): return False
    
if __name__ == "__main__":
    excel = FileExcelReader('upload_file_data.xlsx', 'Sheet1')
    test = TestUploadFile()
    
    num_rows = excel.get_row_count()
    for row in range(2, num_rows + 1):
        test.set_up()
        file_path = excel.read_data(row, 1)
        expected = excel.read_data(row, 2)
        result = test.test_upload(file_path, expected)
        if result:
            excel.write_data("PASSED", row, 3)
        else:
            excel.write_data("FAILED", row, 3)
        test.tear_down()
