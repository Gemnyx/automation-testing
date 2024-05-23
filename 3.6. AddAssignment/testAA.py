import pytest
import re
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.service import Service

@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.maximize_window()
    yield driver
    driver.quit()

def excel_data():
    excel_file = 'AA_datasheet.xlsx'
    sheet_name = 'Sheet1'
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheet_name]
    nRows = sheet.max_row

    data = []
    for row in range(2, nRows + 1):
        name = sheet.cell(row, 2).value
        allowFrom = sheet.cell(row, 3).value
        dueDate = sheet.cell(row, 4).value
        cutOffDate = sheet.cell(row, 5).value
        remindMe = sheet.cell(row, 6).value
        grade = sheet.cell(row, 7).value
        expectedResult = sheet.cell(row, 8).value

        # Process assignment's name
        if name == "null":
            name = ""

        # Process Allow submissions from
        allowFromDay = allowFromMonth = allowFromYear = allowFromHour = allowFromMin = None
        if allowFrom == "null":
            allowFrom = ""
        else:
            splitStr = re.split("\/|:| ", allowFrom)
            allowFromDay = splitStr[0]
            allowFromMonth = splitStr[1]
            allowFromYear = splitStr[2]
            allowFromHour = splitStr[3]
            allowFromMin = splitStr[4]

        # Process Due date
        dueDay = dueMonth = dueYear = dueHour = dueMin = None
        if dueDate == "null":
            dueDate = ""
        else:
            splitStr = re.split("\/|:| ", dueDate)
            dueDay = splitStr[0]
            dueMonth = splitStr[1]
            dueYear = splitStr[2]
            dueHour = splitStr[3]
            dueMin = splitStr[4]

        # Process Cut-off date
        cutOffDay = cutOffMonth = cutOffYear = cutOffHour = cutOffMin = None
        if cutOffDate == "null":
            cutOffDate = ""
        else:
            splitStr = re.split("\/|:| ", cutOffDate)
            cutOffDay = splitStr[0]
            cutOffMonth = splitStr[1]
            cutOffYear = splitStr[2]
            cutOffHour = splitStr[3]
            cutOffMin = splitStr[4]

        # Process Remind me to grade by
        remindDay = remindMonth = remindYear = remindHour = remindMin = None
        if remindMe == "null":
            remindMe = ""
        else:
            splitStr = re.split("\/|:| ", remindMe)
            remindDay = splitStr[0]
            remindMonth = splitStr[1]
            remindYear = splitStr[2]
            remindHour = splitStr[3]
            remindMin = splitStr[4]

        # Process Grade to pass
        if grade == "null":
            grade = ""
                    
        data.append((row, name, allowFromDay, allowFromMonth, allowFromYear, allowFromHour, allowFromMin, dueDay, dueMonth, dueYear, dueHour, dueMin, cutOffDay, cutOffMonth, cutOffYear, cutOffHour, cutOffMin, remindDay, remindMonth, remindYear, remindHour, remindMin, grade, expectedResult))

    workbook.close()
    return data

class TestAddAssignment:
    def update_excel(self, row, result):
        excel_file = 'AA_datasheet.xlsx'
        sheet_name = 'Sheet1'
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=9, value=result)
        workbook.save(excel_file)
        workbook.close()
    
    @pytest.mark.parametrize("row, name, allowFromDay, allowFromMonth, allowFromYear, allowFromHour, allowFromMin, dueDay, dueMonth, dueYear, dueHour, dueMin, cutOffDay, cutOffMonth, cutOffYear, cutOffHour, cutOffMin, remindDay, remindMonth, remindYear, remindHour, remindMin, grade, expectedResult", excel_data())
    def test_addAssignment(self, driver, row, name, allowFromDay, allowFromMonth, allowFromYear, allowFromHour, allowFromMin, dueDay, dueMonth, dueYear, dueHour, dueMin, cutOffDay, cutOffMonth, cutOffYear, cutOffHour, cutOffMin, remindDay, remindMonth, remindYear, remindHour, remindMin, grade, expectedResult):
        # Test logic using the provided WebDriver instance (driver)
        driver.get("https://school.moodledemo.net/login/index.php?loginredirect=1")
        driver.maximize_window()

        # Login as Teacher
        driver.find_element(By.ID, "username").click()
        driver.find_element(By.ID, "username").send_keys("teacher")
        driver.find_element(By.ID, "password").click()
        driver.find_element(By.ID, "password").send_keys("moodle")
        driver.find_element(By.ID, "loginbtn").click()

        # Go to Course
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.LINK_TEXT,"Moodle and Mountaineering")))
        driver.find_element(By.LINK_TEXT,"Moodle and Mountaineering").click()
        link_course = driver.current_url

        # Turn on Edit mode
        element = driver.find_element(By.NAME, "setmode")
        actions = ActionChains(driver)
        actions.move_to_element(element).perform()
        driver.find_element(By.NAME, "setmode").click()

        # Go to Add assignment
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#coursecontentcollapse1 > div.divider.bulk-hidden.d-flex.justify-content-center.align-items-center.always-visible.my-3 > div > button")))
        driver.find_element(By.CSS_SELECTOR, "#coursecontentcollapse1 > div.divider.bulk-hidden.d-flex.justify-content-center.align-items-center.always-visible.my-3 > div > button").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#all-6 > div > div:nth-child(1) > div > a")))
        driver.find_element(By.LINK_TEXT, "#all-6 > div > div:nth-child(1) > div > a").click()
        
        # Data-driven testing - Name
        driver.find_element(By.ID, "id_name").click()
        driver.find_element(By.ID, "id_name").send_keys(name)
        
        # Data-driven testing - Allow submissions from
        if allowFromDay is not None:
            if not driver.find_element(By.ID, "id_allowsubmissionsfromdate_enabled").is_selected():
                driver.find_element(By.ID, "id_allowsubmissionsfromdate_enabled").click()
            selectDay = Select(driver.findElement(By.ID, "id_allowsubmissionsfromdate_day"))
            selectDay.select_by_index(allowFromDay - 1)
            selectMonth = Select(driver.findElement(By.ID, "id_allowsubmissionsfromdate_month"))
            selectMonth.select_by_index(allowFromMonth - 1)
            selectYear = Select(driver.findElement(By.ID, "id_allowsubmissionsfromdate_year"))
            selectYear.select_by_visible_text(allowFromYear)
            selectHour = Select(driver.findElement(By.ID, "id_allowsubmissionsfromdate_hour"))
            selectHour.select_by_visible_text(allowFromHour)
            selectMin = Select(driver.findElement(By.ID, "id_allowsubmissionsfromdate_minute"))
            selectMin.select_by_visible_text(allowFromMin)

        # Data-driven testing - Due date
        if dueDay is not None:
            if not driver.find_element(By.ID, "id_duedate_enabled").is_selected():
                driver.find_element(By.ID, "id_duedate_enabled").click()
            selectDay = Select(driver.findElement(By.ID, "id_duedate_day"))
            selectDay.select_by_index(dueDay - 1)
            selectMonth = Select(driver.findElement(By.ID, "id_duedate_month"))
            selectMonth.select_by_index(dueMonth - 1)
            selectYear = Select(driver.findElement(By.ID, "id_duedate_year"))
            selectYear.select_by_visible_text(dueYear)
            selectHour = Select(driver.findElement(By.ID, "id_duedate_hour"))
            selectHour.select_by_visible_text(dueHour)
            selectMin = Select(driver.findElement(By.ID, "id_duedate_minute"))
            selectMin.select_by_visible_text(dueMin)

        # Data-driven testing - Cut-off date
        if cutOffDay is not None:
            if not driver.find_element(By.ID, "id_cutoffdate_enabled").is_selected():
                driver.find_element(By.ID, "id_cutoffdate_enabled").click()
            selectDay = Select(driver.findElement(By.ID, "id_cutoffdate_day"))
            selectDay.select_by_index(cutOffDay - 1)
            selectMonth = Select(driver.findElement(By.ID, "id_cutoffdate_month"))
            selectMonth.select_by_index(cutOffMonth - 1)
            selectYear = Select(driver.findElement(By.ID, "id_cutoffdate_year"))
            selectYear.select_by_visible_text(cutOffYear)
            selectHour = Select(driver.findElement(By.ID, "id_cutoffdate_hour"))
            selectHour.select_by_visible_text(cutOffHour)
            selectMin = Select(driver.findElement(By.ID, "id_cutoffdate_minute"))
            selectMin.select_by_visible_text(cutOffMin)
        
        # Data-driven testing - Remind me to grade by
        if remindDay is not None:
            if not driver.find_element(By.ID, "id_gradingduedate_enabled").is_selected():
                driver.find_element(By.ID, "id_gradingduedate_enabled").click()
            selectDay = Select(driver.findElement(By.ID, "id_gradingduedate_day"))
            selectDay.select_by_index(remindDay - 1)
            selectMonth = Select(driver.findElement(By.ID, "id_gradingduedate_month"))
            selectMonth.select_by_index(remindMonth - 1)
            selectYear = Select(driver.findElement(By.ID, "id_gradingduedate_year"))
            selectYear.select_by_visible_text(remindYear)
            selectHour = Select(driver.findElement(By.ID, "id_gradingduedate_hour"))
            selectHour.select_by_visible_text(remindHour)
            selectMin = Select(driver.findElement(By.ID, "id_gradingduedate_minute"))
            selectMin.select_by_visible_text(remindMin)

        # Data-driven testing - Grade to pass
        if grade is not None:
            driver.find_element(By.CSS_SELECTOR, "#collapseElement-7").click()
            driver.find_element(By.ID, "id_gradepass").click()
            driver.find_element(By.ID, "id_gradepass").send_keys(grade)
            
            driver.find_element(By.CSS_SELECTOR, "#collapseElement-10").click()
            driver.find_element(By.ID, "id_completion_2").click()
            driver.find_element(By.ID, "id_completionusegrade").click()
            driver.find_element(By.ID, "id_completionpassgrade_1").click()

        # Save and display   
        driver.find_element(By.NAME, "submitbutton2").click()
        
        if expectedResult == "Add new assignment successfully":
            try:
                driver.current_url == link_course
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")
            
        elif expectedResult == "Alert: Invalid name.":
            try:
                driver.find_element(By.ID,'id_error_name')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")
                
        elif expectedResult == "Alert: Due date must be after the allow submissions from date.":
            try:
                driver.find_element(By.ID,'id_error_duedate')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")

        elif expectedResult == "Alert: Cut-off date cannot be earlier than the due date.":
            try:
                driver.find_element(By.ID,'id_error_cutoffdate')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")

        elif expectedResult == "Alert: Remind me to grade by date cannot be earlier than the due date.":
            try:
                driver.find_element(By.ID,'id_error_gradingduedate')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")

        elif expectedResult == "Alert: Cut-off date cannot be earlier than the allow submissions from date.":
            try:
                driver.find_element(By.ID,'id_error_cutoffdate')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")

        driver.find_element(By.ID, "user-menu-toggle").click()
        driver.find_element(By.LINK_TEXT, "Log out").click()