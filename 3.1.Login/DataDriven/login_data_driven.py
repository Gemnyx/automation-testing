import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

from selenium.common.exceptions import NoSuchElementException

class FileExcelReader:
    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet

    def get_row_count(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheet]
        return (sheet.max_row)

    def get_column_count(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheet]
        return (sheet.max_column)

    def read_data(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheet]
        return sheet.cell(row=rownum, column=colnum).value

    def write_data(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheet]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)

class TestLogin:
    def set_up(self):
        self.driver = webdriver.Chrome()
        self.driver.get('https://school.moodledemo.net/login/index.php')
    
    def tear_down(self):
        self.driver.quit()

    def test_login(self, username, password, expected_result):
        time.sleep(0.5)
        self.driver.find_element(By.ID, 'username').send_keys(username)
        time.sleep(1)
        self.driver.find_element(By.ID, 'password').send_keys(password)
        time.sleep(0.5)
        self.driver.find_element(By.ID, 'loginbtn').click()
        time.sleep(1)
        if expected_result == "Successfully Login":
            return True
        elif expected_result == "Invalid Login": 
            errorMessage = self.driver.find_element(By.ID,'loginerrormessage')
            assert errorMessage.text == "Invalid login, please try again"
            return True

if __name__ == "__main__":
    excel = FileExcelReader('login_data.xlsx', 'Sheet1')
    test = TestLogin()
    test.set_up()
    num_rows = excel.get_row_count()
    for row in range(2, num_rows + 1):
        username = excel.read_data(row, 1)
        password = excel.read_data(row, 2)
        expected = excel.read_data(row, 3)
        if username is None:
            username = ""
        if password is None:
            password = ""
        print(f"Testcase {row-1}:\nusername:\t{username}\npassword:\t{password}\nexpected:\t{expected}\nProcessing...")
        try:
            result = test.test_login(username, password, expected)
            excel.write_data("PASSED", row, 4)
            print("Done!\nResult: PASSED\n========================================")
        except:
            excel.write_data("FAILED", row, 4)
            print("Done!\nResult: FAILED\n========================================")
    test.tear_down()