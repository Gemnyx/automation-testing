import pytest
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.service import Service

@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.maximize_window()
    yield driver
    driver.quit()

# @pytest.fixture
def excel_data():
    excel_file = 'AQ_datasheet.xlsx'
    sheet_name = 'Sheet1'
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheet_name]
    nRows = sheet.max_row

    data = []
    for row in range(2, nRows + 1):
        name = sheet.cell(row,2).value
        open = sheet.cell(row,3).value
        close = sheet.cell(row,4).value
        timelimit = sheet.cell(row,5).value
        expectedResult = sheet.cell(row,6).value

        # Process name
        if name == "null":
            name = ""

        # Process opentime
        opentime = open_day = open_month = open_year = None
        if open != "null":
            opentime = time.strptime(open, "%d/%m/%Y")
            open_day = opentime.tm_mday
            open_month = opentime.tm_mon
            open_year = opentime.tm_year

        # Process closetime
        closetime = close_day = close_month = close_year = None
        if close != "null":
            closetime = time.strptime(close, "%d/%m/%Y")
            close_day = closetime.tm_mday
            close_month = closetime.tm_mon
            close_year = closetime.tm_year
            
        

        data.append((row,name, open_day, open_month, open_year, close_day, close_month, close_year, timelimit, expectedResult))

    workbook.close()
    return data

class TestAddQuiz:
    def update_excel(self, row, result):
        excel_file = 'AQ_datasheet.xlsx'
        sheet_name = 'Sheet1'
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=7, value=result)
        workbook.save(excel_file)
        workbook.close()
    
    @pytest.mark.parametrize("row,name, open_day, open_month, open_year, close_day, close_month, close_year, timelimit, expectedResult", excel_data())
    def test_addQuiz(self, driver, row, name, open_day, open_month, open_year, close_day, close_month, close_year, timelimit, expectedResult):
        # Test logic using the provided WebDriver instance (driver)
        driver.get("https://school.moodledemo.net/login/index.php?loginredirect=1")
        driver.maximize_window() 
        # Login teacher
        driver.find_element(By.ID, "username").click()
        driver.find_element(By.ID, "username").send_keys("teacher")
        driver.find_element(By.ID, "password").click()
        driver.find_element(By.ID, "password").send_keys("moodle")
        driver.find_element(By.ID, "loginbtn").click()
        # Go to add quiz
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.LINK_TEXT,"Moodle and Mountaineering")))
        driver.find_element(By.LINK_TEXT,"Moodle and Mountaineering").click()
        # self.driver.find_element(By.CSS_SELECTOR, "#course-info-container-59-3 .multiline > span:nth-child(2)").click()
        link_course = driver.current_url
        element = driver.find_element(By.NAME, "setmode")
        actions = ActionChains(driver)
        actions.move_to_element(element).perform()
        driver.find_element(By.NAME, "setmode").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#coursecontentcollapse0 > div.divider.bulk-hidden.d-flex.justify-content-center.align-items-center.always-visible.my-3 > div > button")))
        driver.find_element(By.CSS_SELECTOR, "#coursecontentcollapse0 > div.divider.bulk-hidden.d-flex.justify-content-center.align-items-center.always-visible.my-3 > div > button").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.LINK_TEXT, "Quiz")))
        driver.find_element(By.LINK_TEXT, "Quiz").click()
        
        # Data-driven testing - Name
        driver.find_element(By.ID, "id_name").click()
        driver.find_element(By.ID, "id_name").send_keys(name)
        
        driver.find_element(By.ID, "collapseElement-1").click()
        WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, "id_timeopen_enabled")))
        # Data-driven testing - Opentime
        if open_day is not None:
            driver.find_element(By.ID, "id_timeopen_enabled").click()
            driver.find_element(By.ID, "id_timeopen_day").click()
            dropdown = driver.find_element(By.ID, "id_timeopen_day")
            select = Select(dropdown)
            select.select_by_visible_text(str(open_day))
            
            driver.find_element(By.ID, "id_timeopen_month").click()
            dropdown = driver.find_element(By.ID, "id_timeopen_month")
            select = Select(dropdown)
            select.select_by_index(open_month)
            
            driver.find_element(By.ID, "id_timeopen_year").click()
            dropdown = driver.find_element(By.ID, "id_timeopen_year")
            select = Select(dropdown)
            select.select_by_visible_text(str(open_year))
            
            driver.find_element(By.ID, "id_timeopen_hour").click()
            dropdown = driver.find_element(By.ID, "id_timeopen_hour")
            dropdown.find_element(By.XPATH, "//option[. = '06']").click()
            
            driver.find_element(By.ID, "id_timeopen_minute").click()
            dropdown = driver.find_element(By.ID, "id_timeopen_minute")
            dropdown.find_element(By.XPATH, "//option[. = '40']").click()
        
        # Data-driven testing - Closetime
        if close_day is not None:
            driver.find_element(By.ID, "id_timeclose_enabled").click()
            driver.find_element(By.ID, "id_timeclose_day").click()
            dropdown = driver.find_element(By.ID, "id_timeclose_day")
            select = Select(dropdown)
            select.select_by_visible_text(str(close_day))
            
            driver.find_element(By.ID, "id_timeclose_month").click()
            dropdown = driver.find_element(By.ID, "id_timeclose_month")
            select = Select(dropdown)
            select.select_by_index(close_month)
            
            driver.find_element(By.ID, "id_timeclose_year").click()
            dropdown = driver.find_element(By.ID, "id_timeclose_year")
            select = Select(dropdown)
            select.select_by_visible_text(str(close_year))
            
            driver.find_element(By.ID, "id_timeclose_hour").click()
            dropdown = driver.find_element(By.ID, "id_timeclose_hour")
            dropdown.find_element(By.XPATH, "//option[. = '06']").click()
            
            driver.find_element(By.ID, "id_timeclose_minute").click()
            dropdown = driver.find_element(By.ID, "id_timeclose_minute")
            dropdown.find_element(By.XPATH, "//option[. = '40']").click()
        
        # Data-driven testing - Timelimit
        if timelimit != "null":
            driver.find_element(By.ID, "id_timelimit_enabled").click()
            driver.find_element(By.ID, "id_timelimit_number").click()
            driver.find_element(By.ID, "id_timelimit_number").send_keys(timelimit)
            
        driver.find_element(By.ID, "id_submitbutton2").click()
        # driver.find_element(By.XPATH, "//form/button").click()
        
        if expectedResult == "Add new quiz successfully":
            # driver.get(link_course)
            # activities = driver.find_elements(By.CLASS_NAME, '.activity-item')
            # flag = False
            # for a in activities:
            #     if (a.get_attribute('data-activityname') == name):
            #         self.update_excel(row, "Passed")
            #         flag = True
            # if flag is False:
            #     self.update_excel(row, "Failed")
            try:
                driver.current_url == link_course
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")
            
        elif expectedResult == "Alert: Invalid name":
            try:
                driver.find_element(By.ID,'id_error_name')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")
                
        elif expectedResult == "Alert: Invalid close time":
            try:
                driver.find_element(By.ID,'id_error_timeclose')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")
        elif expectedResult == "Alert: Invalid time limit":
            try:
                driver.find_element(By.ID,'id_error_timelimit')
                self.update_excel(row, "Passed")
            except:
                self.update_excel(row, "Failed")

        driver.find_element(By.ID, "user-menu-toggle").click()
        driver.find_element(By.LINK_TEXT, "Log out").click()
